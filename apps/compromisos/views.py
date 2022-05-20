from django.shortcuts import render
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.db.models.deletion import ProtectedError
from apps.hallazgos.forms import HallazgoAddForm, HallazgoUpdateForm, HallazgoDetalleForm
from apps.hallazgos.models import Ges_Hallazgo
from apps.auditoria.models import Ges_auditoria, Glo_autoria_auditor
from apps.jefaturas.models import Ges_Jefatura
from apps.compromisos.models import Ges_Compromisos, Ges_Observaciones_Compromiso_Enc

# Create your views here.
from apps.compromisos.forms import CompromisoAddForm, CompromisoResponsableEditForm, \
    CompromisoEditForm, CompromisoEnviarRevisionForm, \
    CompromisoValidaFormAuditor,HallazgoAsignaEncargadoGestion
from django.http import HttpResponseRedirect, JsonResponse
from django.db.models import Q, OuterRef, Count, Subquery

from django.core.mail import EmailMessage,send_mass_mail
from django.contrib.auth.models import User, Group
from django.db.models import Q

from openpyxl import Workbook
from django.http import HttpResponse
from datetime import datetime
# Create your views here.

class HallazgosList(ListView):
    model = Ges_Hallazgo
    template_name = 'compromisos/bandeja_hallazgos_list.html'

    def get_context_data(self,  **kwargs):
        context = super(HallazgosList, self).get_context_data(**kwargs)

        id_usuario = self.request.user.id
        id_jefatura = Ges_Jefatura.objects.get(id_user_id = id_usuario)
        lista_hallazgos = Ges_Hallazgo.objects.filter(Q(id_auditoria_id = self.kwargs['pk'] ) & Q (id_estadoshallazgo_id__in =[ 1,2,3,4,5]))


        context['object_list'] = lista_hallazgos


        #self.request.session['pk_auditoria'] = self.kwargs['pk']
        return context

class ComprimisoList(ListView):
    model = Ges_Compromisos
    template_name = 'compromisos/compromisos_list.html'

    def get_context_data(self,  **kwargs):
        context = super(ComprimisoList, self).get_context_data(**kwargs)

        lista_compromisos = Ges_Compromisos.objects.filter(hallazgo_id=self.kwargs['pk'])
        #auditoria = Ges_auditoria.objects.get(id=self.kwargs['pk'])
        hallazgo =  Ges_Hallazgo.objects.get(id=self.kwargs['pk'])
        context['object_list'] = lista_compromisos
        context['auditoria'] ={ 'nombre' : str(hallazgo.id_auditoria.descripcion_auditoria), 'numero' : str(hallazgo.id_auditoria.cod_auditoria) }
        context['hallazgo'] = {'id':hallazgo.id_auditoria_id, 'estado_hallazgo':hallazgo.id_estadoshallazgo_id}
        self.request.session['pk_hallazgo'] = self.kwargs['pk']
        return context


class CompromisoCreate(SuccessMessageMixin, CreateView ):
    form_class = CompromisoAddForm
    template_name = 'compromisos/compromisos_form.html'

    def post(self, request, *args, **kwargs):
        form_class = self.get_form_class()
        form = self.get_form(form_class)

        print(form.errors.as_json())
        if form.is_valid():

            form.instance.hallazgo_id_id = self.request.session['pk_hallazgo']
            form.instance.estado_compromiso_id= '1'
            form.save()
            actualizaEstadoHallazgo(2, self.request.session['pk_hallazgo'])
            request.session['message_class'] = "alert alert-success"
            messages.success(self.request, "Los datos fueron creados correctamente!")
            return HttpResponseRedirect('/compromisos/listarCompromiso/'+ str(self.request.session['pk_hallazgo']))
        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha creado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/compromisos/listarCompromiso/'+ str(self.request.session['pk_hallazgo']))

class CompromisoCreateEg(SuccessMessageMixin, CreateView ):
    form_class = CompromisoAddForm
    template_name = 'compromisos/compromisos_form.html'

    def post(self, request, *args, **kwargs):
        form_class = self.get_form_class()
        form = self.get_form(form_class)

        print(form.errors.as_json())
        if form.is_valid():

            form.instance.hallazgo_id_id = self.request.session['pk_hallazgo']
            form.instance.estado_compromiso_id= '1'
            form.save()
            actualizaEstadoHallazgo(2, self.request.session['pk_hallazgo'])
            request.session['message_class'] = "alert alert-success"
            messages.success(self.request, "Los datos fueron creados correctamente!")
            return HttpResponseRedirect('/bandeja_gestion/listarCompromisoEnc/'+ str(self.request.session['pk_hallazgo']))
        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha creado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/bandeja_gestion/listarCompromisoEnc/'+ str(self.request.session['pk_hallazgo']))


class HallazgoCreateEncargadoGestion(SuccessMessageMixin, UpdateView ):
    model = Ges_Hallazgo
    form_class = HallazgoAsignaEncargadoGestion
    template_name = 'compromisos/hallazgo_reasignar_eg_form.html'

    def get_context_data(self,  **kwargs):
        context = super(HallazgoCreateEncargadoGestion, self).get_context_data(**kwargs)
        try:
            hallazgo = Ges_Hallazgo.objects.get(id=self.kwargs['pk'])
        except Ges_Hallazgo.DoesNotExist:
            return None

        self.request.session['enc_gestion_id_anterior'] = str(hallazgo.enc_gestion_id_id)


        context['hallazgo'] = {'null':'',}

        return context

    def post(self, request, *args, **kwargs):

        self.object = self.get_object
        id = kwargs['pk']
        id_hallazgo = self.model.objects.get(id=id)
        form = self.form_class(request.POST, instance=id_hallazgo)



        # print(form.errors.as_json())
        if form.is_valid():



            id_user = request.POST['enc_gestion_id']  # aquí capturo lo que traigo del modal
            id_user_anterior=self.request.session['enc_gestion_id_anterior']

            if id_user_anterior == 'None':
                id_user_anterior = 0

            cant_enc = Ges_Hallazgo.objects.filter(enc_gestion_id_id=id_user_anterior).count()

            en_grupo = list(Group.objects.values_list('user', flat=True).filter(id=4))


            # Esta validación reemplaza al usuario del grupo si posee solo 1 hallazgo asignado, de lo contrario
            # se agrega, lo cual evita dejar usuarios que no posean hallazgos, en el grupo de encargados

            if (id_user not in str(en_grupo)):
                if (id_user_anterior == 'None'):
                    g = Group.objects.get(id=4)
                    g.user_set.add(id_user)
                else:
                    g = Group.objects.get(id=4)
                    g.user_set.add(id_user)
                    if cant_enc == 1:
                        g.user_set.remove(id_user_anterior)

            form.instance.id_estadoshallazgo_id = '3'

            form.save()

            request.session['message_class'] = "alert alert-success"
            messages.success(self.request, "Los datos fueron creados correctamente!")
            return HttpResponseRedirect('/compromisos/listarHallazgoBandeja/')
        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha creado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/compromisos/listarHallazgoBandeja/')


def actualizaEstadoHallazgo(estado, id_hallazgo):


    Ges_Hallazgo.objects.filter(id=id_hallazgo).update(

        id_estadoshallazgo_id=estado,
    )
    return None

class CompromisoDelete(SuccessMessageMixin, DeleteView ):
    model = Ges_Compromisos
    template_name = 'compromisos/compromisos_delete.html'

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        try:
            obj.delete()


            request.session['message_class'] = "alert alert-success"
            messages.success(self.request, "El registro fue eliminado correctamente!")
            return HttpResponseRedirect('/compromisos/listarCompromiso/'+ str(self.request.session['pk_hallazgo']))

        except ProtectedError as e:
            request.session['message_class'] = "alert alert-danger"
            messages.error(request, "Error de integridad: No se puede eliminar este compromiso porque tiene registros asociados.")
            return HttpResponseRedirect('/compromisos/listarCompromiso/'+ str(self.request.session['pk_hallazgo']))

        except Exception  as e:
            request.session['message_class'] = "alert alert-danger"
            messages.error(request, "Error interno: No se ha eliminado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/compromisos/listarCompromiso/'+ str(self.request.session['pk_hallazgo']))

class CompromisoDeleteEg(SuccessMessageMixin, DeleteView ):
    model = Ges_Compromisos
    template_name = 'compromisos/compromisos_delete.html'

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        try:
            obj.delete()


            request.session['message_class'] = "alert alert-success"
            messages.success(self.request, "El registro fue eliminado correctamente!")
            return HttpResponseRedirect('/bandeja_gestion/listarCompromisoEnc/'+ str(self.request.session['pk_hallazgo']))

        except ProtectedError as e:
            request.session['message_class'] = "alert alert-danger"
            messages.error(request, "Error de integridad: No se puede eliminar este compromiso porque tiene registros asociados.")
            return HttpResponseRedirect('/bandeja_gestion/listarCompromisoEnc/'+ str(self.request.session['pk_hallazgo']))

        except Exception  as e:
            request.session['message_class'] = "alert alert-danger"
            messages.error(request, "Error interno: No se ha eliminado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/bandeja_gestion/listarCompromisoEnc/'+ str(self.request.session['pk_hallazgo']))




class CompromisoUpdate(SuccessMessageMixin, UpdateView ):
    model = Ges_Compromisos
    form_class = CompromisoEditForm
    template_name = 'compromisos/compromisos_form.html'


    def get_context_data(self,  **kwargs):
        context = super(CompromisoUpdate, self).get_context_data(**kwargs)
        try:
            compromiso = Ges_Compromisos.objects.get(id=self.kwargs['pk'])
        except Ges_Compromisos.DoesNotExist:
            return None


        context['compromisos'] = {'estado':str(compromiso.estado_compromiso_id),}

        return context

    def post(self, request, *args, **kwargs):

        self.object = self.get_object
        id = kwargs['pk']
        id_compromiso = self.model.objects.get(id=id)
        form = self.form_class(request.POST, instance=id_compromiso)

        if form.is_valid():
            form.save()
            request.session['message_class'] = "alert alert-success"
            messages.success(self.request, "Los datos fueron actualizados correctamente!" )
            return HttpResponseRedirect('/compromisos/listarCompromiso/'+ str(self.request.session['pk_hallazgo']))
        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha actualizado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/compromisos/listarCompromiso/' + str(self.request.session['pk_hallazgo']))


class CompromisoUpdateAuditor(SuccessMessageMixin, UpdateView ):
    model = Ges_Compromisos
    form_class = CompromisoEditForm
    template_name = 'compromisos/compromisos_form.html'

    #
    # def get_context_data(self,  **kwargs):
    #     context = super(CompromisoUpdateAuditor, self).get_context_data(**kwargs)
    #     try:
    #         compromiso = Ges_Compromisos.objects.get(id=self.kwargs['pk'])
    #     except Ges_Compromisos.DoesNotExist:
    #         return None
    #
    #
    #     context['compromisos'] = {'estado':str(compromiso.estado_compromiso_id),}
    #
    #     return context

    def post(self, request, *args, **kwargs):

        self.object = self.get_object
        id = kwargs['pk']
        id_compromiso = self.model.objects.get(id=id)
        form = self.form_class(request.POST, instance=id_compromiso)

        if form.is_valid():
            form.save()
            request.session['message_class'] = "alert alert-success"
            messages.success(self.request, "Los datos fueron actualizados correctamente!" )
            return HttpResponseRedirect('/compromisos/hallazgoDetalleCompromisoAuditor/' + str(self.request.session['pk_hallazgo']))
        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha actualizado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/compromisos/hallazgoDetalleCompromisoAuditor/' + str(self.request.session['pk_hallazgo']))

class CompromisoUpdateEg(SuccessMessageMixin, UpdateView ):
    model = Ges_Compromisos
    form_class = CompromisoEditForm
    template_name = 'compromisos/compromisos_form.html'


    def get_context_data(self,  **kwargs):
        context = super(CompromisoUpdateEg, self).get_context_data(**kwargs)
        try:
            compromiso = Ges_Compromisos.objects.get(id=self.kwargs['pk'])
        except Ges_Compromisos.DoesNotExist:
            return None


        context['compromisos'] = {'estado':str(compromiso.estado_compromiso_id),}

        return context

    def post(self, request, *args, **kwargs):

        self.object = self.get_object
        id = kwargs['pk']
        id_compromiso = self.model.objects.get(id=id)
        form = self.form_class(request.POST, instance=id_compromiso)

        if form.is_valid():
            form.save()
            request.session['message_class'] = "alert alert-success"
            messages.success(self.request, "Los datos fueron actualizados correctamente!" )
            return HttpResponseRedirect('/bandeja_gestion/listarCompromisoEnc/'+ str(self.request.session['pk_hallazgo']))
        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha actualizado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/bandeja_gestion/listarCompromisoEnc/' + str(self.request.session['pk_hallazgo']))



class CompromisoValidarAuditor(SuccessMessageMixin, UpdateView ):
    model = Ges_Compromisos
    form_class = CompromisoValidaFormAuditor
    template_name = 'compromisos/compromiso_form_valida_auditor.html'


    def get_context_data(self,  **kwargs):
        context = super(CompromisoValidarAuditor, self).get_context_data(**kwargs)
        try:
            compromiso = Ges_Compromisos.objects.get(id=self.kwargs['pk'])
        except Ges_Compromisos.DoesNotExist:
            return None


        context['compromisos'] = {'id':str(compromiso.id) }




        return context

    def post(self, request, *args, **kwargs):

        self.object = self.get_object
        id = kwargs['pk']
        id_compromiso = self.model.objects.get(id=id)
        form = self.form_class(request.POST, instance=id_compromiso)

        if form.is_valid():
            form.save()
            request.session['message_class'] = "alert alert-success"
            messages.success(self.request, "Los datos fueron actualizados correctamente!" )
            return HttpResponseRedirect('/compromisos/listarCompromiso/'+ str(self.request.session['pk_hallazgo']))
        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha actualizado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/compromisos/listarCompromiso/' + str(self.request.session['pk_hallazgo']))

class CompromisoResponsableUpdate(SuccessMessageMixin, UpdateView ):
    model = Ges_Compromisos
    form_class = CompromisoResponsableEditForm
    template_name = 'compromisos/compromiso_responsable_form.html'

    def post(self, request, *args, **kwargs):

        self.object = self.get_object
        id = kwargs['pk']
        id_compromiso = self.model.objects.get(id=id)
        form = self.form_class(request.POST, instance=id_compromiso)

        if form.is_valid():
            form.save()
            request.session['message_class'] = "alert alert-success"
            messages.success(self.request, "Los datos fueron actualizados correctamente!" )
            return HttpResponseRedirect('/compromisos/listarCompromiso/'+ str(self.request.session['pk_hallazgo']))
        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha actualizado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/compromisos/listarCompromiso/' + str(self.request.session['pk_hallazgo']))


class CompromisoResponsableUpdateEg(SuccessMessageMixin, UpdateView ):
    model = Ges_Compromisos
    form_class = CompromisoResponsableEditForm
    template_name = 'compromisos/compromiso_responsable_form.html'

    def post(self, request, *args, **kwargs):

        self.object = self.get_object
        id = kwargs['pk']
        id_compromiso = self.model.objects.get(id=id)
        form = self.form_class(request.POST, instance=id_compromiso)

        if form.is_valid():
            form.save()
            request.session['message_class'] = "alert alert-success"
            messages.success(self.request, "Los datos fueron actualizados correctamente!" )
            return HttpResponseRedirect('/bandeja_gestion/listarCompromisoEnc/'+ str(self.request.session['pk_hallazgo']))
        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha actualizado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/bandeja_gestion/listarCompromisoEnc/' + str(self.request.session['pk_hallazgo']))

class EnviaCompromiso(UpdateView): # #########################################################
    model = Ges_Compromisos
    form_class = CompromisoEnviarRevisionForm
    template_name = 'compromisos/compromiso_form_envia_ra.html'


    def get_context_data(self,  **kwargs):
        context = super(EnviaCompromiso, self).get_context_data(**kwargs)

        try:
            id_compromiso = Ges_Compromisos.objects.get(id=self.kwargs['pk'])
        except Ges_Compromisos.DoesNotExist:
            return None

        # context['hallazgo'] = {'id': id_compromiso.id}



        context = {"tiene_responsable": id_compromiso.responsable_hallazgo_id}

        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id = kwargs['pk']
        id_compromiso = self.model.objects.get(id=id)
        # id_usuario_actual = self.request.user.id
        form = self.form_class(request.POST, instance=id_compromiso)

        id_compromiso= Ges_Compromisos.objects.get(id=id)

        id_hallazgo= Ges_Hallazgo.objects.get(id=id_compromiso.hallazgo_id_id)

        id_auditoria= id_hallazgo.id_auditoria



        auditores = list(Glo_autoria_auditor.objects.values_list('id_auditor_id', flat=True).filter(id_auditoria_id=id_auditoria))
        auditores_emails = list(
            User.objects.values_list('email', flat=True).filter(id__in=auditores))




        try:

            form.instance.estado_compromiso_id = '2'
            form.save()
            actualizaEstadoHallazgo(5, id_compromiso.hallazgo_id_id)

        except:

            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request,
                           "Error interno: El compromiso no pudo ser enviado por un error interno. Comuníquese con el administrador.")
            return HttpResponseRedirect('/compromisos/listarCompromiso/' + str(self.request.session['pk_hallazgo']))


        try:

            EnviarCorreoInicioSeguimiento(auditores_emails, id_auditoria.cod_auditoria,
                                          id_auditoria.descripcion_auditoria)
            request.session['message_class'] = "alert alert-success"  # Tipo mensaje
            messages.success(request,
                             "El compromiso fue enviado correctamente y el correo enviado a los auditores!")  # mensaje
            return HttpResponseRedirect('/compromisos/listarCompromiso/' + str(self.request.session['pk_hallazgo']))  # Redirije a la pantalla principal

        except:

            request.session['message_class'] = "alert alert-warning"  # Tipo mensaje
            messages.success(request,
                             "El compromiso fue enviado correctamente!, pero el servicio de correo tuvo un inconveniente favor comuníquese con la jefatura directa para informar del envío.")  # mensaje
            return HttpResponseRedirect('/compromisos/listarCompromiso/' + str(self.request.session['pk_hallazgo']))  # Redirije a la pantalla principal



class CompromisoDevolverEG(UpdateView):
    model = Ges_Compromisos
    form_class = CompromisoEnviarRevisionForm
    template_name = 'compromisos/compromiso_devolver_eg.html'


    # def get_context_data(self,  **kwargs):
    #     context = super(CompromisoDevolverEG, self).get_context_data(**kwargs)
    #
    #     try:
    #         id_compromiso = Ges_Compromisos.objects.get(id=self.kwargs['pk'])
    #     except Ges_Hallazgo.DoesNotExist:
    #         return None
    #
    #     # context['hallazgo'] = {'id': id_compromiso.id}
    #
    #
    #
    #     context = {"tiene_responsable": id_compromiso.responsable_hallazgo_id}
    #
    #     return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id = kwargs['pk']
        id_compromiso = self.model.objects.get(id=id)
        # id_usuario_actual = self.request.user.id
        form = self.form_class(request.POST, instance=id_compromiso)

        id_compromiso= Ges_Compromisos.objects.get(id=id)

        id_hallazgo= Ges_Hallazgo.objects.get(id=id_compromiso.hallazgo_id_id)

        id_auditoria= id_hallazgo.id_auditoria


        #
        # auditores = list(Glo_autoria_auditor.objects.values_list('id_auditor_id', flat=True).filter(id_auditoria_id=id_auditoria))
        # auditores_emails = list(
        #     User.objects.values_list('email', flat=True).filter(id__in=auditores))


        try:

            form.instance.estado_compromiso_id = '6'
            form.save()
            actualizaEstadoHallazgo(3,id_compromiso.hallazgo_id_id)

        except:

            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request,
                           "Error interno: El compromiso no pudo ser enviado por un error interno. Comuníquese con el administrador.")
            return HttpResponseRedirect('/compromisos/listarCompromiso/' + str(self.request.session['pk_hallazgo']))


        try:

            # EnviarCorreoInicioSeguimiento(auditores_emails, id_auditoria.cod_auditoria,
            #                               id_auditoria.descripcion_auditoria)
            request.session['message_class'] = "alert alert-success"  # Tipo mensaje
            messages.success(request,
                             "El compromiso fue devuelto correctamente y el correo enviado al encargado de gestión!")  # mensaje
            return HttpResponseRedirect('/compromisos/listarCompromiso/' + str(self.request.session['pk_hallazgo']))  # Redirije a la pantalla principal

        except:

            request.session['message_class'] = "alert alert-warning"  # Tipo mensaje
            messages.success(request,
                             "El compromiso fue devuelto correctamente!, pero el servicio de correo tuvo un inconveniente favor comuníquese con el encargado de gestión para informar del envío.")  # mensaje
            return HttpResponseRedirect('/compromisos/listarCompromiso/' + str(self.request.session['pk_hallazgo']))  # Redirije a la pantalla principal

class hallazgoDetalleCompomiso(SuccessMessageMixin, UpdateView ):
    model = Ges_Hallazgo
    form_class = HallazgoDetalleForm
    template_name = 'compromisos/hallazgo_detalle_compromiso.html'

    def get_context_data(self,  **kwargs):
        context = super(hallazgoDetalleCompomiso, self).get_context_data(**kwargs)

        try:
            hallazgo = Ges_Hallazgo.objects.get(id=self.kwargs['pk'])
        except Ges_Hallazgo.DoesNotExist:
            return None

        context['hallazgo'] = {'id': hallazgo.id}

        return context

def update_compromiso_rechaza(request):
    id_compromiso = request.POST.get('id')
    response_data = {}
    if request.POST.get('action') == 'post':
        #Valores a recibir del form
        id_compromiso = request.POST.get('id_compromiso')
        comentario_auditor = request.POST.get('comentario_auditor')
        estado_rechazado = request.POST.get('estado_rechazado')


        #buscar el modelo a actualizar
        compromiso = Ges_Compromisos.objects.get(id=id_compromiso)
        compromiso.comentario_auditor  = comentario_auditor
        compromiso.estado_compromiso_id = estado_rechazado

        compromiso_obj = Ges_Compromisos.objects.get(id=id_compromiso)

        hallazgo = Ges_Hallazgo.objects.get(id=compromiso_obj.hallazgo_id_id)

        id_auditoria= hallazgo.id_auditoria

        jefatura_emails = list(
            User.objects.values_list('email', flat=True).filter(id=hallazgo.jefatura_id.id_user_id))

        #
        try:
          compromiso.save()

          EnviarCorreoRechazaCompromiso(jefatura_emails, id_auditoria.cod_auditoria,
                                       id_auditoria.descripcion_auditoria)



          response_data['error'] = 'El compromiso fue rechazado correctamente.'
        except:
          response_data['error'] = 'Error al intentar validar el compromiso, intente nuevamente o comuniquese con el administrador.'

        return JsonResponse(response_data)

    return render(request, 'compromisos/hallazgoDetalleCompromisoAuditor/' + str(id_compromiso), {'success': 'true'})



def update_compromiso_acepta(request):
    id_compromiso = request.POST.get('id')
    response_data = {}
    if request.POST.get('action') == 'post':
        #Valores a recibir del form
        id_compromiso = request.POST.get('id_compromiso')
        comentario_auditor = request.POST.get('comentario_auditor')
        estado_aceptado = request.POST.get('estado_aceptado')
        estado_hallazgo_aceptado= request.POST.get('estado_hallazgo_aceptado')


        #buscar el modelo a actualizar
        compromiso = Ges_Compromisos.objects.get(id=id_compromiso)
        compromiso.comentario_auditor  = comentario_auditor
        compromiso.estado_compromiso_id = estado_aceptado







        compromiso_obj = Ges_Compromisos.objects.get(id=id_compromiso)



        hallazgo = Ges_Hallazgo.objects.get(id=compromiso_obj.hallazgo_id_id)



        id_auditoria= hallazgo.id_auditoria

        jefatura_emails = list(
            User.objects.values_list('email', flat=True).filter(id=hallazgo.jefatura_id.id_user_id))


        #
        try:



          #actualiza el estado del compromiso a "Validado por Auditor"
          compromiso.save()
          #actualiza el estado del hallazgo a "Validado por Auditor"
          Ges_Hallazgo.objects.filter(id=compromiso.hallazgo_id_id).update(id_estadoshallazgo_id=estado_hallazgo_aceptado)




          EnviarCorreoValidaCompromiso(jefatura_emails, id_auditoria.cod_auditoria,
                                        id_auditoria.descripcion_auditoria)

          response_data['error'] = 'El compromiso fue aceptado correctamente.'
        except:
          response_data['error'] = 'Error al intentar validar el compromiso, intente nuevamente o comuniquese con el administrador.'

        return JsonResponse(response_data)

    return render(request, 'compromisos/hallazgoDetalleCompromisoAuditor/' + str(id_compromiso), {'success': 'true'})


def EnviarCorreoInicioSeguimiento(auditores_emails, cod_auditoria, descripcion_auditoria):

    # ahora = datetime.now()
    # fecha = ahora.strftime("%d" + "/" + "%m" + "/" + "%Y" + " a las " + "%H:%M")
    cod_auditoria = str(cod_auditoria)
    descripcion_auditoria = str(descripcion_auditoria)
    idcorreoJefatura=[auditores_emails]



    subject = 'Revisión Compromiso de Auditoría'
    messageHtml = '<b>Estimado/a Auditor/a</b>: <br><br> En el marco de la auditoria en curso <b>'+ cod_auditoria +'</b> <b>'+ descripcion_auditoria +'</b>, se requiere su ingreso al Sistema de Auditoria Institucional, para proceder a la revisión del/los compromisos ingresados por la contraparte. <b> <br> El link de acceso es:  <a href="http://seguimientoauditoria.ine.cl:8008/accounts/login/"> Sistema Auditoría </a> </b> <br> <br>Saludos cordiales, <br>Sistema de Auditoria. <br><p style="font-size:12px;color:red;">correo generado automaticamente favor no responder.'

    email = EmailMessage(subject, messageHtml ,to=idcorreoJefatura)
    email.content_subtype='html'
    email.send()



def EnviarCorreoValidaCompromiso(responsable_hallazgo, cod_auditoria, descripcion_auditoria):

    # ahora = datetime.now()
    # fecha = ahora.strftime("%d" + "/" + "%m" + "/" + "%Y" + " a las " + "%H:%M")
    cod_auditoria = str(cod_auditoria)
    descripcion_auditoria = str(descripcion_auditoria)
    idcorreoJefatura=[responsable_hallazgo]



    subject = 'Revisión Compromiso de Auditoría'
    messageHtml = '<b>Estimado/a</b>: <br><br> En el marco de la auditoria en curso <b>'+ cod_auditoria +'</b> <b>'+ descripcion_auditoria +'</b>, realizada por el Depto de Auditoria Institucional, se informa que su compromiso se encuentra validado. <b> <br> El link de acceso es:  <a href="http://seguimientoauditoria.ine.cl:8008/accounts/login/"> Sistema Auditoría </a> </b><br><br> <br>Agradeciendo desde ya su gestión y participación <br>Saludos cordiales, <br>Depto de Auditoria Institucional <br><p style="font-size:12px;color:red;">correo generado automaticamente favor no responder.'

    email = EmailMessage(subject, messageHtml ,to=idcorreoJefatura)
    email.content_subtype='html'
    email.send()


def EnviarCorreoRechazaCompromiso(responsable_hallazgo, cod_auditoria, descripcion_auditoria):

    # ahora = datetime.now()
    # fecha = ahora.strftime("%d" + "/" + "%m" + "/" + "%Y" + " a las " + "%H:%M")
    cod_auditoria = str(cod_auditoria)
    descripcion_auditoria = str(descripcion_auditoria)
    idcorreoJefatura=[responsable_hallazgo]


    subject = 'Revisión Compromiso de Auditoría'
    messageHtml = '<b>Estimado/a</b>: <br><br> En el marco de la auditoria en curso <b>'+ cod_auditoria +'</b> <b>'+ descripcion_auditoria +'</b>, realizada por el Depto de Auditoria Institucional,  se informa que su compromiso ha sido observado, por lo cual deberá ingresar a sistema y proceder según corresponda. <b> <br> El link de acceso es:  <a href="http://seguimientoauditoria.ine.cl:8008/accounts/login/ "> Sistema Auditoría </a> </b><br><br> <br>Agradeciendo desde ya su gestión y participación <br>Saludos cordiales, <br>Depto de Auditoria Institucional <br><p style="font-size:12px;color:red;">correo generado automaticamente favor no responder.'

    email = EmailMessage(subject, messageHtml ,to=idcorreoJefatura)
    email.content_subtype='html'
    email.send()

class ComprimisoAuditorList(ListView):
    model = Ges_Compromisos
    template_name = 'compromisos/hallazgo_compromiso_auditor_list.html'

    def get_context_data(self,  **kwargs):
        context = super(ComprimisoAuditorList, self).get_context_data(**kwargs)

        lista_compromisos = Ges_Compromisos.objects.filter(hallazgo_id=self.kwargs['pk'])
        hallazgo =  Ges_Hallazgo.objects.get(id=self.kwargs['pk'])
        context['object_list'] = lista_compromisos
        context['auditoria'] ={ 'nombre' : str(hallazgo.id_auditoria.descripcion_auditoria), 'numero' : str(hallazgo.id_auditoria.cod_auditoria) }

        context['id_hallazgo'] = {'id': str(hallazgo.id_auditoria_id)}

        # context = {"id_hallazgo": hallazgo.id}



        self.request.session['pk_hallazgo'] = self.kwargs['pk']
        return context


class ComprimisoAuditorListDirector(ListView):
    model = Ges_Compromisos
    template_name = 'compromisos/hallazgo_compromiso_auditor_list_director.html'

    def get_context_data(self,  **kwargs):
        context = super(ComprimisoAuditorListDirector, self).get_context_data(**kwargs)


        count_observaciones_compromiso = Ges_Observaciones_Compromiso_Enc.objects.values('compromiso_observacion_id').filter(
            compromiso_observacion_id=OuterRef('pk')).annotate(
            count_observaciones_compromiso=Count('id'))

        lista_compromisos = Ges_Compromisos.objects.filter(hallazgo_id=self.kwargs['pk']).annotate(
            count_observaciones=Subquery(count_observaciones_compromiso.values('count_observaciones_compromiso')))


        hallazgo =  Ges_Hallazgo.objects.get(id=self.kwargs['pk'])
        context['object_list'] = lista_compromisos
        context['auditoria'] ={ 'nombre' : str(hallazgo.id_auditoria.descripcion_auditoria), 'numero' : str(hallazgo.id_auditoria.cod_auditoria) }

        context['id_hallazgo'] = {'id': str(hallazgo.id_auditoria_id)}

        # context = {"id_hallazgo": hallazgo.id}



        self.request.session['pk_hallazgo'] = self.kwargs['pk']
        return context


def export_users_xls_auditoria_auditor(request, *args, **kwargs):

    id_auditoria = kwargs['pk']

    # auditoria= Ges_auditoria.objects.filter(id=id_auditoria)
    # hallazgo=Ges_Hallazgo.objects.filter(id_auditoria=id_auditoria)
    compromiso=Ges_Compromisos.objects.filter(hallazgo_id__id_auditoria=id_auditoria)




    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Reporte_Auditoria_Hallazgo.xlsx'.format(
        date=datetime.now().strftime('%d/%m/%Y'),
    )
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'reporte_auditoria'


    columns = ['N°',
               'Proceso/Subproceso/Etapa/Proyecto/Otro',
               'Naturaleza del Trabajo',
               'Descripción del Hallazgo (Condición)',
               'Criterios',
               'Causas',
               'Efectos',
               'Sumario',
               'Recomendación',
               'Plazo',
               'Responsable',
               'Encargado Gestión (si existe)',
               'Descripción Compromiso',
               'Medio de Verificación',
               'Plazo Implementación',
               'Responsable',
               'Cargo Responsable'


               ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for compromiso in compromiso:
        row_num += 1
        row = ''

        if (compromiso.responsable_hallazgo is None):
            validadaStr = 'Valor no ingresado'
        else:
            validadaStr = compromiso.responsable_hallazgo.id_user

        if (compromiso.cargo_responsable_id is None):
            CargoStr = 'Valor no ingresado'
        else:
            CargoStr = compromiso.cargo_responsable_id


        row = [compromiso.hallazgo_id_id,
               str(compromiso.hallazgo_id.proceso),
               str(compromiso.hallazgo_id.naturaleza),
               str(compromiso.hallazgo_id.descripcion_hallazgo),
               str(compromiso.hallazgo_id.criterios),
               str(compromiso.hallazgo_id.causas),
               str(compromiso.hallazgo_id.efectos),
               str(compromiso.hallazgo_id.sumario),
               str(compromiso.hallazgo_id.recomendacion),
               compromiso.hallazgo_id.plazo,
               str(compromiso.hallazgo_id.jefatura_id.id_user),
               str(compromiso.hallazgo_id.enc_gestion_id),
               str(compromiso.descripcion_compromiso),
               str(compromiso.medio_verificacion),
               compromiso.plazo_implementacion,
               str(validadaStr),
               str(CargoStr),




               ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value


            if col_num == 10:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 15:
                cell.number_format = 'dd/mm/yyyy'
            # if col_num == 14:
            #     cell.number_format = 'dd/mm/yyyy'
            # if col_num == 15:
            #     cell.number_format = 'dd/mm/yyyy'
            # if col_num == 16:
            #     cell.number_format = 'dd/mm/yyyy'
            # if col_num == 17:
            #     cell.number_format = 'dd/mm/yyyy:HH:MM:SS'

    workbook.save(response)


    return response


def export_users_xls_auditoria_auditado(request, *args, **kwargs):

    id_user = kwargs['pk']

    # auditoria= Ges_auditoria.objects.filter(id=id_auditoria)
    # hallazgo=Ges_Hallazgo.objects.filter(id_auditoria=id_auditoria)

    id_jefatura= Ges_Jefatura.objects.get(id_user_id=id_user)


    compromiso=Ges_Compromisos.objects.filter(hallazgo_id__jefatura_id=id_jefatura.id)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Reporte_Auditoria_Hallazgo.xlsx'.format(
        date=datetime.now().strftime('%d/%m/%Y'),
    )
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'reporte_auditoria'


    columns = ['N°',
               'Proceso/Subproceso/Etapa/Proyecto/Otro',
               'Naturaleza del Trabajo',
               'Descripción del Hallazgo (Condición)',
               'Criterios',
               'Causas',
               'Efectos',
               'Recomendación',
               'Plazo',
               'Responsable',
               'Encargado Gestión (si existe)',
               'Descripción Compromiso',
               'Medio de Verificación',
               'Plazo Implementación',
               'Responsable',
               'Cargo Responsable',
               'Estado Hallazgo'


               ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for compromiso in compromiso:
        row_num += 1
        row = ''

        if (compromiso.responsable_hallazgo is None):
            validadaStr = 'Valor no ingresado'
        else:
            validadaStr = compromiso.responsable_hallazgo.id_user

        if (compromiso.cargo_responsable_id is None):
            CargoStr = 'Valor no ingresado'
        else:
            CargoStr = compromiso.cargo_responsable_id


        row = [compromiso.hallazgo_id_id,
               str(compromiso.hallazgo_id.proceso),
               str(compromiso.hallazgo_id.naturaleza),
               str(compromiso.hallazgo_id.descripcion_hallazgo),
               str(compromiso.hallazgo_id.criterios),
               str(compromiso.hallazgo_id.causas),
               str(compromiso.hallazgo_id.efectos),
               str(compromiso.hallazgo_id.recomendacion),
               compromiso.hallazgo_id.plazo,
               str(compromiso.hallazgo_id.jefatura_id.id_user),
               str(compromiso.hallazgo_id.enc_gestion_id),
               str(compromiso.descripcion_compromiso),
               str(compromiso.medio_verificacion),
               compromiso.plazo_implementacion,

               str(validadaStr),
               str(CargoStr),
               str(compromiso.hallazgo_id.id_estadoshallazgo)




               ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value


            if col_num == 10:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 15:
                cell.number_format = 'dd/mm/yyyy'
            # if col_num == 14:
            #     cell.number_format = 'dd/mm/yyyy'
            # if col_num == 15:
            #     cell.number_format = 'dd/mm/yyyy'
            # if col_num == 16:
            #     cell.number_format = 'dd/mm/yyyy'
            # if col_num == 17:
            #     cell.number_format = 'dd/mm/yyyy:HH:MM:SS'

    workbook.save(response)


    return response
